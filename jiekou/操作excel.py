import xlrd
import os
from openpyxl import workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import xlwt
from xlutils.copy import copy
nempaht = os.chdir(r'C:\token')
filename = 'rippton测试用例.xlsx'
file = os.path.join(os.getcwd(),filename)


xl = xlrd.open_workbook(file)
#print(xl.sheet_names())
#print(xl.sheets())
#print(xl.nsheets)
#print(xl.sheet_names('11111'))
#print(xl.sheet_by_index(0))

tabel = xl.sheet_by_name('目录')
#print(tabel.name)
#print(tabel.nrows)列
#print(tabel.ncols)行
#print(tabel.row_values(0,1,4))
#print(tabel.row(0))
#print(tabel.row_types(0))
#print(tabel.col_values(0,0,2))
#print(tabel.row_slice(0))
#print(tabel.row_types(0,1,2))

#wbk = xlwt.Workbook()
#sheet = wbk.add_sheet('sheet 1')
#sheet.write(0,1,'test text')
#wbk.save('test.xlsx')


#workbook = xlrd.open_workbook(file)
#workbooknew=copy(workbook)
#ws = workbooknew.get_sheet(0)
#ws.write(3,0,'changed!')
#workbooknew.save(file)

#print(tabel.cell_value(1,2))
#print(tabel.cell(1,2).value)
#print(tabel.row(1)[2].value)
#print(tabel.cell_type(1,2))
#print(xlrd.cellname(0,0))
#print(xlrd.colname(30))
#wbk = xlwt.Workbook()
#sheet = wbk.add_sheet('你是我的小苹果')
#sheet.write(0,1,'我爱你你是我的朱丽叶')
#wbk.save('一个人.xls')


#workbook = xlrd.open_workbook('一个人.xls')
#workbooknew=copy(workbook)
#ws = workbooknew.get_sheet(0)
#ws.write(0,0,'这才是第一个')
#workbooknew.save('yigerencopyban.xls')
